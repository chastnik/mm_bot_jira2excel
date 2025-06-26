from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from typing import List, Dict
import logging
from datetime import datetime
import io

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Генератор Excel отчетов с трудозатратами"""
    
    def __init__(self):
        # Порядок колонок согласно шаблону
        self.column_order = [
            "date",         # Дата работы
            "executor",     # Исполнитель
            "hours",        # Часы
            "description",  # Содержание работы
            "project_task", # Проектная задача
            "project"       # Проект
        ]
    
    def generate_timesheet_report(self, worklogs: List[Dict], project_name: str, start_date: str, end_date: str, projects: List[Dict] = None) -> bytes:
        """
        Генерировать Excel отчет с трудозатратами согласно шаблону
        
        Args:
            worklogs: Список данных о трудозатратах
            project_name: Название проекта
            start_date: Дата начала периода
            end_date: Дата окончания периода
            projects: Список проектов (для сводного отчета)
            
        Returns:
            Байты Excel файла
        """
        try:
            # Создаем новую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = f"Трудозатраты {project_name}"
            
            # Добавляем строку заголовков (первая строка) начиная с колонки A
            ws.cell(row=1, column=1, value="Дата работы")       # A1
            ws.cell(row=1, column=2, value="Исполнитель")       # B1  
            ws.cell(row=1, column=3, value="Часы")              # C1
            ws.cell(row=1, column=4, value="Содержание работы") # D1
            ws.cell(row=1, column=5, value="Проектная задача")  # E1
            ws.cell(row=1, column=6, value="Проект")            # F1
            
            # Заполняем данными начиная со второй строки
            for row, worklog in enumerate(worklogs, 2):  # Начинаем со второй строки
                # Парсим дату из строки и записываем как текст в формате DD.MM.YYYY HH:MM:SS
                try:
                    # Парсим дату в формате "2025-6-18 14:30"
                    date_obj = datetime.strptime(worklog['date'], '%Y-%m-%d %H:%M')
                    # Форматируем как текст в российском формате
                    formatted_date = date_obj.strftime('%d.%m.%Y %H:%M:%S')
                    date_cell = ws.cell(row=row, column=1, value=formatted_date)
                    # Явно устанавливаем формат ячейки как "Общий" (текст)
                    date_cell.number_format = 'General'
                except ValueError:
                    # Если не удалось распарсить, записываем как есть
                    date_cell = ws.cell(row=row, column=1, value=worklog['date'])
                    date_cell.number_format = 'General'
                
                ws.cell(row=row, column=2, value=worklog['executor'])    # B - Исполнитель
                ws.cell(row=row, column=3, value=worklog['hours'])       # C - Часы
                ws.cell(row=row, column=4, value=worklog['description']) # D - Содержание работы
                ws.cell(row=row, column=5, value=worklog['project_task'])# E - Проектная задача
                ws.cell(row=row, column=6, value=worklog['project'])     # F - Проект
            
            # Автоширина столбцов
            for col in range(1, 7):  # A-F
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = self._get_column_width(col)
            
            # Сохраняем в память
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Сгенерирован Excel отчет по шаблону с заголовками и {len(worklogs)} записями")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Ошибка генерации Excel отчета: {e}")
            raise
    
    def _get_column_width(self, column_index: int) -> int:
        """Получить оптимальную ширину столбца"""
        widths = {
            1: 20,  # A - Дата работы
            2: 15,  # B - Исполнитель
            3: 10,  # C - Часы
            4: 50,  # D - Содержание работы
            5: 25,  # E - Проектная задача
            6: 20   # F - Проект
        }
        return widths.get(column_index, 15)
    
    def generate_filename(self, project_name: str, start_date: str, end_date: str) -> str:
        """Генерировать имя файла для отчета"""
        # Убираем недопустимые символы из названия проекта
        safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_project_name = safe_project_name.replace(' ', '_')
        
        return f"trudozatraty_{safe_project_name}_{start_date}_{end_date}.xlsx"
    
    def generate_filename_for_multiple_projects(self, projects: List[Dict], start_date: str, end_date: str) -> str:
        """Генерировать имя файла для отчета по нескольким проектам"""
        if len(projects) == 1:
            return self.generate_filename(projects[0]['name'], start_date, end_date)
        
        # Для нескольких проектов используем их ключи
        project_keys = [p['key'] for p in projects]
        
        # Если проектов слишком много, используем сокращенное имя
        if len(project_keys) > 3:
            filename_projects = f"{len(project_keys)}_proektov"
        else:
            filename_projects = "_".join(project_keys)
        
        return f"trudozatraty_svodnyj_{filename_projects}_{start_date}_{end_date}.xlsx" 